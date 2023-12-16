import smtplib
from email.message import EmailMessage
import sys
import openpyxl
import re


def email_check(email_id):
    pattern = re.compile(
        r"^[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*@(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?$")
    check = pattern.fullmatch(email_id) #check if entered data matches above regular expression
    if check:
        print('Entered email id is valid \nsending mail...')
        send_mail(email_id) 
        return
    else:
        print('Invalid Email id')
        return

def send_mail(mail_id):
    email = EmailMessage() # inserting contents of email using email object
    email['from'] = 'sender\'s_name' #Enter name of sender here
    email['to'] = mail_id
    email['subject'] = 'subject' #Enter subject here
    email.set_content('contents') #Enter contents of mail here


    with smtplib.SMTP(host='smtp.gmail.com', port=587) as smtp: # creates smtp client session object
        smtp.ehlo()
        smtp.starttls() # upgrade to secure connection
        #Enter mail id and password or passkey below
        smtp.login('mail_id', 'password/passkey') # login to your mail account
        
        smtp.send_message(email) #send email
        print('mail sent!!!')
    return

def main(file):
    email_list = openpyxl.load_workbook(file) # load excel file as object
    sheet = email_list.active # select all active columns
    while True:
        try:
            row = int(input('Enter row number : ')) # enter row number where title is located
            column = int(input('Enter column number : ')) # enter column number where title is located
            # make sure your data is in vertical format otherwise program will terminate
            10/row
            10/column
        except ValueError as err:
            print(f'please enter a number \n {err}')
        except ZeroDivisionError as err:
            print(f'please enter a number greater than zero \n {err}')
        else:
            for i in range(sheet.max_row - 1):
                row += 1 # skips title at first loop
                mail_id = sheet.cell(column = column, row = row) # creating cell object in order to access its values     
                email_check(str(mail_id.value))
            sys.exit()

if __name__ == '__main__':
    main(sys.argv[1])
    